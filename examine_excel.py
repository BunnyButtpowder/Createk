#!/usr/bin/env python3
"""
Examine the structure of the Createk product Excel file
"""

import pandas as pd
import openpyxl
import sys
import io

# Set UTF-8 encoding for stdout
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

excel_file = 'public/2026 _ Danh sách sản phẩm Createk.xlsx'

# Load with openpyxl to see all sheets
wb = openpyxl.load_workbook(excel_file)

print("Available sheets:")
for sheet_name in wb.sheetnames:
    print(f"  - {sheet_name}")

# Check product codes across sheets
print("Checking product code mapping between sheets...")
print(f"{'='*60}")

# Get product codes from detail sheet
df_details = pd.read_excel(excel_file, sheet_name="Danh sách SP (Chi tiết)")
df_details.columns = df_details.iloc[0]
df_details = df_details.drop(0).reset_index(drop=True)
df_details.columns = df_details.columns.str.strip()

print(f"\nProduct codes from 'Danh sách SP (Chi tiết)' (first 10):")
print(f"Available columns: {list(df_details.columns)}")
# Find the column that might contain product codes
code_col = None
for col in df_details.columns:
    if 'AT' in str(col) or 'code' in str(col).lower() or 'mã' in str(col).lower():
        code_col = col
        print(f"Found potential code column: {col}")
        break

if code_col:
    detail_codes = df_details[code_col].dropna().unique().tolist()
    for code in detail_codes[:10]:
        print(f"  {code}")
else:
    print("No product code column found")
    print("First few rows of the dataframe:")
    print(df_details.head(3).to_string())

# Get product codes from URL sheet
df_urls = pd.read_excel(excel_file, sheet_name="Mã SP chi tiết")
df_urls.columns = df_urls.iloc[0]
df_urls = df_urls.drop(0).reset_index(drop=True)
df_urls.columns = df_urls.columns.str.strip()

print(f"\nProduct codes from 'Mã SP chi tiết' (first 10):")
print(f"Available columns: {list(df_urls.columns)}")
# Find the column that might contain product codes
code_col_url = None
for col in df_urls.columns:
    if 'AT' in str(col) or 'code' in str(col).lower() or 'mã' in str(col).lower():
        code_col_url = col
        print(f"Found potential code column: {col}")
        break

if code_col_url:
    url_codes = df_urls[code_col_url].dropna().unique().tolist()
    for code in url_codes[:10]:
        print(f"  {code}")
else:
    print("No product code column found")
    print("First few rows of the dataframe:")
    print(df_urls.head(3).to_string())

# Check for overlap
if code_col and code_col_url:
    overlap = set(detail_codes) & set(url_codes)
    print(f"\nOverlapping product codes: {len(overlap)}")
    print(f"  Examples: {list(overlap)[:5] if overlap else 'None'}")

    # Check if there's a pattern difference
    print(f"\nSample detailed products with codes:")
    for _, row in df_details.head(5).iterrows():
        code = str(row.get(code_col, ''))
        name_col = None
        for col in df_details.columns:
            if 'sản phẩm' in str(col).lower() or 'product' in str(col).lower() or 'tên' in str(col).lower():
                name_col = col
                break
        name = str(row.get(name_col, ''))[:50] if name_col else ''
        print(f"  Code: {code:15} | Name: {name}")

    print(f"\nSample URL products with codes:")
    for _, row in df_urls.head(5).iterrows():
        code = str(row.get(code_col_url, ''))
        url_col = None
        for col in df_urls.columns:
            if 'ảnh' in str(col).lower() or 'image' in str(col).lower() or 'link' in str(col).lower():
                url_col = col
                break
        url = str(row.get(url_col, ''))[:60] if url_col and pd.notna(row.get(url_col)) else ''
        print(f"  Code: {code:15} | URL: {url}")
else:
    print("Cannot check overlap - missing code columns")

print(f"\n{'='*60}")
print("Now examining all sheets structure...")
print(f"{'='*60}")

# Read the two target sheets plus check other sheets for links
sheets_to_read = [
    "Danh sách sản phẩm (Nhóm to)",
    "Danh sách SP (Chi tiết)",
    "Mã SP chi tiết"
]

for sheet_name in sheets_to_read:
    if sheet_name in wb.sheetnames:
        print(f"\n{'='*60}")
        print(f"Sheet: {sheet_name}")
        print(f"{'='*60}")

        df = pd.read_excel(excel_file, sheet_name=sheet_name)
        print(f"Shape: {df.shape[0]} rows × {df.shape[1]} columns")
        print(f"\nColumns:")
        for i, col in enumerate(df.columns, 1):
            print(f"  {i}. {col}")

        print(f"\nFirst few rows:")
        print(df.head(3).to_string())

        print(f"\nData types:")
        print(df.dtypes.to_string())

        # Look for URLs and links
        print(f"\nSearching for URLs/links...")
        urls = []
        for col in df.columns:
            for idx, val in df[col].items():
                if pd.notna(val) and isinstance(val, str) and ('http' in val or 'www.' in val):
                    urls.append({'row': idx, 'column': col, 'value': val[:150]})

        if urls:
            print(f"Found {len(urls)} URLs:")
            for url in urls[:15]:  # Show first 15
                print(f"  Row {url['row']}, Col '{url['column']}': {url['value']}")
        else:
            print("No URLs found")

        # Show sample data from key columns
        if sheet_name == "Danh sách sản phẩm (Nhóm to)":
            print(f"\nShowing 'Link ảnh' column content (row 1-10):")
            if 'Unnamed: 4' in df.columns:  # Link ảnh column
                for idx, val in df['Unnamed: 4'].head(10).items():
                    if pd.notna(val):
                        print(f"  Row {idx}: {val}")

        if sheet_name == "Danh sách SP (Chi tiết)":
            print(f"\nShowing 'Mã Createk (Link drive tổng)' column content (row 1-10):")
            if 'Mã Createk \n(Link drive tổng)' in df.columns:
                for idx, val in df['Mã Createk \n(Link drive tổng)'].head(10).items():
                    if pd.notna(val):
                        print(f"  Row {idx}: {val}")
    else:
        print(f"\nSheet '{sheet_name}' not found!")
